#!/usr/bin/env python3
"""
Agenda Renta4 - Load Sections from Excel
Importa las URLs desde el archivo Excel original a la tabla 'sections'
"""

import sys
import os
from pathlib import Path
import sqlite3

try:
    import openpyxl
except ImportError:
    print("❌ ERROR: openpyxl no está instalado")
    print("Instalar con: pip install openpyxl")
    sys.exit(1)


DATABASE_PATH = os.getenv('DATABASE_PATH', 'agendaRenta4.db')
EXCEL_PATH = Path(__file__).parent / "original-data" / "251028_Árbol web - control calidad.xlsx"


def generate_section_name(url, hierarchy_levels):
    """
    Genera un nombre descriptivo para la sección a partir de la URL.

    Args:
        url: URL completa (ej: "https://www.r4.com/planes-de-pensiones/categorias")
        hierarchy_levels: Lista de niveles de jerarquía (Col 2-6)

    Returns:
        Nombre descriptivo (ej: "Planes de Pensiones - Categorías")
    """
    if not url or not isinstance(url, str):
        return None

    # Intenta extraer de URL
    try:
        # Obtener path sin dominio
        if 'r4.com' in url:
            path = url.split('r4.com')[-1]
        else:
            path = url

        # Dividir por / y tomar últimos 2 segmentos
        parts = [p for p in path.split('/') if p]
        if len(parts) >= 2:
            # Convertir guiones a espacios y capitalizar
            name_parts = []
            for part in parts[-2:]:
                words = part.replace('-', ' ').replace('_', ' ').title()
                name_parts.append(words)
            return ' - '.join(name_parts)
        elif len(parts) == 1:
            return parts[0].replace('-', ' ').replace('_', ' ').title()
    except:
        pass

    # Fallback: usar jerarquía
    hierarchy_names = [h for h in hierarchy_levels if h and isinstance(h, str) and h.strip()]
    if hierarchy_names:
        # Tomar últimos 2 niveles
        return ' - '.join(hierarchy_names[-2:]) if len(hierarchy_names) >= 2 else hierarchy_names[-1]

    # Fallback final: usar URL completa
    return url


def load_sections_from_excel(excel_path, db_path):
    """
    Lee el Excel y carga las secciones (URLs) en la base de datos.

    Args:
        excel_path: Ruta al archivo Excel
        db_path: Ruta a la base de datos SQLite
    """
    print(f"📊 Cargando secciones desde: {excel_path}\n")

    # Verificar que el archivo existe
    if not Path(excel_path).exists():
        print(f"❌ ERROR: Archivo no encontrado: {excel_path}")
        sys.exit(1)

    # Cargar workbook
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"❌ ERROR al cargar Excel: {e}")
        sys.exit(1)

    # Obtener pestaña principal
    if "Actualización y calidad" not in wb.sheetnames:
        print(f"❌ ERROR: Pestaña 'Actualización y calidad' no encontrada")
        print(f"   Pestañas disponibles: {wb.sheetnames}")
        sys.exit(1)

    sheet = wb["Actualización y calidad"]

    # Conectar a BD
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Estadísticas
    total_rows = 0
    inserted = 0
    skipped = 0
    errors = 0

    print("📝 Procesando filas...\n")

    # Limitar a primeras 200 filas (suficiente para capturar todas las URLs reales)
    MAX_ROWS = 200

    # Iterar sobre filas (saltando header)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=MAX_ROWS, values_only=True), start=2):
        total_rows += 1

        # Columna 7 (index 6) = URL
        url = row[6] if len(row) > 6 else None

        # Filtrar URLs vacías o inválidas
        if not url or not isinstance(url, str) or not url.strip():
            skipped += 1
            continue

        # Filtrar si no es URL válida (debe contener r4.com o empezar con /)
        if 'r4.com' not in url and not url.startswith('/'):
            skipped += 1
            continue

        # Obtener jerarquía (Cols 2-6, index 1-5)
        hierarchy = [row[i] if len(row) > i else None for i in range(1, 6)]

        # Generar nombre descriptivo
        name = generate_section_name(url, hierarchy)
        if not name:
            name = url  # Fallback

        # Insertar en BD
        try:
            cursor.execute("""
                INSERT INTO sections (name, url, active)
                VALUES (?, ?, 1)
            """, (name, url))
            inserted += 1
            print(f"   ✓ {name}")
            print(f"     {url}")
        except sqlite3.IntegrityError:
            # URL duplicada (ya existe)
            skipped += 1
        except Exception as e:
            errors += 1
            print(f"   ✗ Error en fila {row_idx}: {e}")

    conn.commit()
    conn.close()

    # Resumen
    print("\n" + "=" * 80)
    print(f"📊 RESUMEN:")
    print(f"   - Filas procesadas: {total_rows}")
    print(f"   - Secciones insertadas: {inserted}")
    print(f"   - Filas omitidas: {skipped}")
    print(f"   - Errores: {errors}")
    print("=" * 80 + "\n")

    if inserted == 0:
        print("⚠️  No se insertaron secciones. Revisa el archivo Excel.")
    else:
        print(f"✅ {inserted} secciones cargadas correctamente\n")


def main():
    """
    Script principal.
    """
    # Verificar que la BD existe
    if not Path(DATABASE_PATH).exists():
        print(f"❌ ERROR: Base de datos no encontrada: {DATABASE_PATH}")
        print(f"   Ejecuta primero: python database.py")
        sys.exit(1)

    # Cargar secciones
    load_sections_from_excel(EXCEL_PATH, DATABASE_PATH)

    # Mostrar estadísticas finales
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM sections WHERE active = 1")
    count = cursor.fetchone()[0]
    conn.close()

    print(f"📊 Total de secciones activas en BD: {count}\n")


if __name__ == '__main__':
    main()