#!/usr/bin/env python3
"""
Generate Excel Report from Crawl Results
Creates comprehensive Excel file with all discovered URLs and statistics
"""

import os
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils import db_cursor

load_dotenv()

def get_crawl_stats(crawl_run_id):
    """Get statistics from latest crawl run"""
    with db_cursor() as cursor:
        # Get crawl run info
        cursor.execute("""
            SELECT
                id,
                started_at,
                finished_at,
                status,
                urls_discovered,
                created_by,
                EXTRACT(EPOCH FROM (finished_at - started_at)) as duration_seconds
            FROM crawl_runs
            WHERE id = %s
        """, (crawl_run_id,))

        crawl_run = cursor.fetchone()

        # Get URLs by depth
        cursor.execute("""
            SELECT depth, COUNT(*) as count
            FROM discovered_urls
            WHERE crawl_run_id = %s
            GROUP BY depth
            ORDER BY depth
        """, (crawl_run_id,))

        depth_stats = cursor.fetchall()

        return crawl_run, depth_stats

def get_all_discovered_urls(crawl_run_id):
    """Get all discovered URLs"""
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT
                u.id,
                u.url,
                u.depth,
                u.discovered_at,
                u.is_priority,
                u.status_code,
                u.response_time,
                u.is_broken,
                u.error_message,
                u.last_checked,
                parent.url as parent_url
            FROM discovered_urls u
            LEFT JOIN discovered_urls parent ON u.parent_url_id = parent.id
            WHERE u.crawl_run_id = %s
            ORDER BY u.id
        """, (crawl_run_id,))

        return cursor.fetchall()

def get_hardcoded_urls():
    """Get hardcoded URLs from sections table"""
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT id, name, url
            FROM sections
            WHERE active = TRUE
            ORDER BY id
        """)

        return cursor.fetchall()

def compare_urls(discovered_urls, hardcoded_urls):
    """Compare discovered vs hardcoded URLs"""
    discovered_set = set(row['url'] for row in discovered_urls)
    hardcoded_set = set(row['url'] for row in hardcoded_urls)

    matches = discovered_set & hardcoded_set
    missing = hardcoded_set - discovered_set
    new = discovered_set - hardcoded_set

    return matches, missing, new

def create_excel_report(crawl_run_id):
    """Create comprehensive Excel report"""

    print("=" * 80)
    print("GENERATING EXCEL REPORT")
    print("=" * 80)

    # Get data
    print("\n1. Fetching crawl statistics...")
    crawl_run, depth_stats = get_crawl_stats(crawl_run_id)

    print("2. Fetching all discovered URLs...")
    discovered_urls = get_all_discovered_urls(crawl_run_id)

    print("3. Fetching hardcoded URLs...")
    hardcoded_urls = get_hardcoded_urls()

    print("4. Comparing URLs...")
    matches, missing, new = compare_urls(discovered_urls, hardcoded_urls)

    # Create workbook
    print("5. Creating Excel workbook...")
    wb = Workbook()

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === SHEET 1: RESUMEN EJECUTIVO ===
    print("6. Creating 'Resumen' sheet...")
    ws_summary = wb.active
    ws_summary.title = "Resumen"

    # Title
    ws_summary['A1'] = "INFORME DE CRAWL - R4.COM"
    ws_summary['A1'].font = Font(size=18, bold=True, color="366092")
    ws_summary.merge_cells('A1:D1')

    # Date
    ws_summary['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_summary['A2'].font = Font(italic=True, size=10)

    # Crawl info
    row = 4
    ws_summary[f'A{row}'] = "INFORMACIÓN DEL CRAWL"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    info_data = [
        ["Crawl Run ID:", crawl_run['id']],
        ["Estado:", crawl_run['status'].upper()],
        ["Iniciado:", crawl_run['started_at'].strftime('%d/%m/%Y %H:%M:%S')],
        ["Finalizado:", crawl_run['finished_at'].strftime('%d/%m/%Y %H:%M:%S')],
        ["Duración:", f"{int(crawl_run['duration_seconds'] / 60)} minutos"],
        ["Creado por:", crawl_run['created_by']],
    ]

    for label, value in info_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1

    # Statistics
    row += 2
    ws_summary[f'A{row}'] = "ESTADÍSTICAS GENERALES"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    # Count priority URLs
    priority_count = sum(1 for url in discovered_urls if url['is_priority'])
    non_priority_count = len(discovered_urls) - priority_count

    stats_data = [
        ["URLs Descubiertas:", len(discovered_urls)],
        ["  - URLs Prioritarias (auditadas):", priority_count],
        ["  - URLs Nuevas (descubiertas):", non_priority_count],
        ["URLs en Lista Manual:", len(hardcoded_urls)],
        ["URLs Coincidentes:", len(matches)],
        ["URLs Faltantes (no encontradas):", len(missing)],
        ["URLs Nuevas:", len(new)],
        ["Cobertura:", f"{(len(matches) / len(hardcoded_urls) * 100):.1f}%"],
    ]

    for label, value in stats_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1

    # Depth statistics
    row += 2
    ws_summary[f'A{row}'] = "DISTRIBUCIÓN POR PROFUNDIDAD"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    ws_summary[f'A{row}'] = "Nivel"
    ws_summary[f'B{row}'] = "URLs"
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary[f'B{row}'].fill = header_fill
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'B{row}'].font = header_font
    row += 1

    for depth_row in depth_stats:
        ws_summary[f'A{row}'] = f"Profundidad {depth_row['depth']}"
        ws_summary[f'B{row}'] = depth_row['count']
        row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 30

    # === SHEET 2: TODAS LAS URLs ===
    print("7. Creating 'Todas las URLs' sheet...")
    ws_all = wb.create_sheet("Todas las URLs")

    headers = ["#", "URL", "⭐", "Estado", "Código", "Tiempo(s)", "Profundidad", "Descubierta", "URL Padre"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(1, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, url_row in enumerate(discovered_urls, 2):
        ws_all.cell(idx, 1).value = idx - 1
        ws_all.cell(idx, 2).value = url_row['url']
        ws_all.cell(idx, 3).value = "⭐" if url_row['is_priority'] else ""

        # Status
        if url_row['last_checked']:
            if url_row['is_broken']:
                ws_all.cell(idx, 4).value = "❌ Roto"
                ws_all.cell(idx, 4).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            else:
                ws_all.cell(idx, 4).value = "✅ OK"
                ws_all.cell(idx, 4).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        else:
            ws_all.cell(idx, 4).value = "⚪ No validada"

        # Status code
        ws_all.cell(idx, 5).value = url_row['status_code'] if url_row['status_code'] else '-'

        # Response time
        if url_row['response_time']:
            ws_all.cell(idx, 6).value = f"{url_row['response_time']:.2f}"
        else:
            ws_all.cell(idx, 6).value = '-'

        ws_all.cell(idx, 7).value = url_row['depth']
        ws_all.cell(idx, 8).value = url_row['discovered_at'].strftime('%d/%m/%Y %H:%M:%S') if url_row['discovered_at'] else '-'
        ws_all.cell(idx, 9).value = url_row['parent_url'] or '-'

        # Highlight priority URLs
        if url_row['is_priority']:
            ws_all.cell(idx, 3).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws_all.cell(idx, 3).font = Font(bold=True, size=14)

        # Borders
        for col in range(1, 10):
            ws_all.cell(idx, col).border = border

    # Adjust column widths
    ws_all.column_dimensions['A'].width = 8
    ws_all.column_dimensions['B'].width = 70
    ws_all.column_dimensions['C'].width = 5
    ws_all.column_dimensions['D'].width = 15
    ws_all.column_dimensions['E'].width = 10
    ws_all.column_dimensions['F'].width = 12
    ws_all.column_dimensions['G'].width = 12
    ws_all.column_dimensions['H'].width = 20
    ws_all.column_dimensions['I'].width = 70

    # Freeze panes
    ws_all.freeze_panes = 'A2'

    # Auto filter
    ws_all.auto_filter.ref = f"A1:I{len(discovered_urls) + 1}"

    # === SHEET 3: URLs PRIORITARIAS ===
    print("8. Creating 'URLs Prioritarias' sheet...")
    ws_priority = wb.create_sheet("URLs Prioritarias")

    headers = ["#", "URL", "Estado", "Código", "Tiempo(s)", "Profundidad", "Descubierta"]
    for col, header in enumerate(headers, 1):
        cell = ws_priority.cell(1, col)
        cell.value = header
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Filter priority URLs
    priority_urls = [url for url in discovered_urls if url['is_priority']]

    for idx, url_row in enumerate(priority_urls, 2):
        ws_priority.cell(idx, 1).value = idx - 1
        ws_priority.cell(idx, 2).value = url_row['url']

        # Status
        if url_row['last_checked']:
            if url_row['is_broken']:
                ws_priority.cell(idx, 3).value = "❌ Roto"
                ws_priority.cell(idx, 3).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            else:
                ws_priority.cell(idx, 3).value = "✅ OK"
                ws_priority.cell(idx, 3).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        else:
            ws_priority.cell(idx, 3).value = "⚪ No validada"

        ws_priority.cell(idx, 4).value = url_row['status_code'] if url_row['status_code'] else '-'

        if url_row['response_time']:
            ws_priority.cell(idx, 5).value = f"{url_row['response_time']:.2f}"
        else:
            ws_priority.cell(idx, 5).value = '-'

        ws_priority.cell(idx, 6).value = url_row['depth']
        ws_priority.cell(idx, 7).value = url_row['discovered_at'].strftime('%d/%m/%Y %H:%M:%S') if url_row['discovered_at'] else '-'

        for col in range(1, 8):
            ws_priority.cell(idx, col).border = border

    ws_priority.column_dimensions['A'].width = 8
    ws_priority.column_dimensions['B'].width = 80
    ws_priority.column_dimensions['C'].width = 15
    ws_priority.column_dimensions['D'].width = 10
    ws_priority.column_dimensions['E'].width = 12
    ws_priority.column_dimensions['F'].width = 12
    ws_priority.column_dimensions['G'].width = 20
    ws_priority.freeze_panes = 'A2'
    ws_priority.auto_filter.ref = f"A1:G{len(priority_urls) + 1}"

    # === SHEET 4: URLs COINCIDENTES ===
    print("9. Creating 'URLs Coincidentes' sheet...")
    ws_matches = wb.create_sheet("URLs Coincidentes")

    headers = ["#", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws_matches.cell(1, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for idx, url in enumerate(sorted(matches), 2):
        ws_matches.cell(idx, 1).value = idx - 1
        ws_matches.cell(idx, 2).value = url

        for col in range(1, 3):
            ws_matches.cell(idx, col).border = border

    ws_matches.column_dimensions['A'].width = 8
    ws_matches.column_dimensions['B'].width = 100
    ws_matches.freeze_panes = 'A2'

    # === SHEET 5: URLs FALTANTES ===
    print("10. Creating 'URLs Faltantes' sheet...")
    ws_missing = wb.create_sheet("URLs Faltantes")

    headers = ["#", "URL", "Nombre en Lista"]
    for col, header in enumerate(headers, 1):
        cell = ws_missing.cell(1, col)
        cell.value = header
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.border = border

    missing_with_names = [(h['url'], h['name']) for h in hardcoded_urls if h['url'] in missing]

    for idx, (url, name) in enumerate(sorted(missing_with_names), 2):
        ws_missing.cell(idx, 1).value = idx - 1
        ws_missing.cell(idx, 2).value = url
        ws_missing.cell(idx, 3).value = name

        for col in range(1, 4):
            ws_missing.cell(idx, col).border = border

    ws_missing.column_dimensions['A'].width = 8
    ws_missing.column_dimensions['B'].width = 80
    ws_missing.column_dimensions['C'].width = 40
    ws_missing.freeze_panes = 'A2'

    # === SHEET 6: URLs NUEVAS ===
    print("11. Creating 'URLs Nuevas' sheet...")
    ws_new = wb.create_sheet("URLs Nuevas")

    headers = ["#", "URL"]
    for col, header in enumerate(headers, 1):
        cell = ws_new.cell(1, col)
        cell.value = header
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.border = border

    for idx, url in enumerate(sorted(new), 2):
        ws_new.cell(idx, 1).value = idx - 1
        ws_new.cell(idx, 2).value = url

        for col in range(1, 3):
            ws_new.cell(idx, col).border = border

    ws_new.column_dimensions['A'].width = 8
    ws_new.column_dimensions['B'].width = 100
    ws_new.freeze_panes = 'A2'

    # Save workbook
    filename = f"informe_crawl_r4_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"\n12. Saving Excel file: {filename}")
    wb.save(filename)

    print(f"\n✅ Excel report created successfully: {filename}")
    return filename

def create_csv_backup(crawl_run_id):
    """Create CSV backup file"""
    print("\n13. Creating CSV backup...")

    discovered_urls = get_all_discovered_urls(crawl_run_id)

    filename = f"urls_todas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    with open(filename, 'w', encoding='utf-8') as f:
        f.write("URL,Profundidad,Descubierta,URL_Padre\n")
        for url_row in discovered_urls:
            discovered_at = url_row['discovered_at'].strftime('%Y-%m-%d %H:%M:%S') if url_row['discovered_at'] else '-'
            parent = url_row['parent_url'] or '-'
            f.write(f'"{url_row["url"]}",{url_row["depth"]},{discovered_at},"{parent}"\n')

    print(f"✅ CSV backup created: {filename}")
    return filename

def create_txt_backup(crawl_run_id):
    """Create plain text backup file"""
    print("\n14. Creating TXT backup...")

    discovered_urls = get_all_discovered_urls(crawl_run_id)

    filename = f"urls_todas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    with open(filename, 'w', encoding='utf-8') as f:
        for url_row in discovered_urls:
            f.write(f"{url_row['url']}\n")

    print(f"✅ TXT backup created: {filename}")
    return filename

if __name__ == '__main__':
    print("\n📊 GENERADOR DE INFORME EXCEL - CRAWL R4.COM\n")

    # Get latest crawl run ID
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT id
            FROM crawl_runs
            WHERE created_by = 'full-crawl-comparison'
            ORDER BY id DESC
            LIMIT 1
        """)
        result = cursor.fetchone()

        if not result:
            print("❌ No crawl run found")
            exit(1)

        crawl_run_id = result['id']

    print(f"Using crawl_run_id: {crawl_run_id}\n")

    # Generate reports
    excel_file = create_excel_report(crawl_run_id)
    csv_file = create_csv_backup(crawl_run_id)
    txt_file = create_txt_backup(crawl_run_id)

    print("\n" + "=" * 80)
    print("✅ TODOS LOS ARCHIVOS GENERADOS EXITOSAMENTE")
    print("=" * 80)
    print(f"\n📄 Archivos creados:")
    print(f"   1. {excel_file} - Informe completo con múltiples hojas")
    print(f"   2. {csv_file} - Respaldo en formato CSV")
    print(f"   3. {txt_file} - Lista simple de URLs")
    print("\n✨ Listo para compartir con tu mujer!")
    print("=" * 80 + "\n")
