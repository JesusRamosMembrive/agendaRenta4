#!/usr/bin/env python3
"""
Script de exploración del Excel original.

Objetivo: Entender estructura de las pestañas antes de diseñar BD.
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no está instalado")
    print("Instalar con: pip install openpyxl")
    sys.exit(1)


def explore_excel(file_path):
    """
    Explora estructura del archivo Excel y muestra información detallada.
    """
    print(f"📊 Explorando: {file_path}\n")
    print("=" * 80)

    # Cargar workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"❌ ERROR: Archivo no encontrado: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ ERROR al cargar Excel: {e}")
        sys.exit(1)

    # Listar todas las pestañas
    print("\n📑 PESTAÑAS ENCONTRADAS:")
    print("-" * 80)
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"{i}. {sheet_name}")

    # Explorar cada pestaña
    for sheet_name in wb.sheetnames:
        print("\n" + "=" * 80)
        print(f"\n📋 PESTAÑA: {sheet_name}")
        print("=" * 80)

        sheet = wb[sheet_name]

        # Dimensiones
        print(f"\n📐 Dimensiones:")
        print(f"   - Filas con datos: {sheet.max_row}")
        print(f"   - Columnas con datos: {sheet.max_column}")

        # Headers (primera fila)
        print(f"\n📊 COLUMNAS (Primera fila como headers):")
        print("-" * 80)
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col_idx)
            header = cell.value if cell.value else f"[Vacía_{col_idx}]"
            headers.append(header)
            print(f"   Col {col_idx:2d}: {header}")

        # Primeras 3 filas de datos (para entender contenido)
        print(f"\n📝 PRIMERAS 3 FILAS DE DATOS:")
        print("-" * 80)

        max_rows_to_show = min(4, sheet.max_row + 1)  # Filas 2-3 (saltando header)
        for row_idx in range(2, max_rows_to_show):
            print(f"\n   Fila {row_idx}:")
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                # Truncar valores muy largos
                if isinstance(value, str) and len(value) > 60:
                    value = value[:57] + "..."
                print(f"      {header}: {value}")

        # Estadísticas de columnas (limitado a primeras 50 filas)
        print(f"\n📈 ESTADÍSTICAS POR COLUMNA (primeras 50 filas):")
        print("-" * 80)

        max_rows_to_analyze = min(50, sheet.max_row + 1)
        for col_idx, header in enumerate(headers, 1):
            # Contar celdas no vacías
            non_empty = 0
            unique_values = set()

            for row_idx in range(2, max_rows_to_analyze):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and cell.value != "":
                    non_empty += 1
                    # Guardar valores únicos (máximo 10 para no saturar)
                    if len(unique_values) < 10:
                        unique_values.add(str(cell.value)[:50])  # Truncar

            print(f"\n   {header}:")
            print(f"      - Celdas con datos: {non_empty}/{max_rows_to_analyze - 2} (muestra)")
            print(f"      - Valores únicos (muestra): {len(unique_values)}")
            if unique_values and len(unique_values) <= 10:
                print(f"      - Ejemplos: {sorted(unique_values)}")

    print("\n" + "=" * 80)
    print("\n✅ Exploración completada")
    print("\n💡 Próximo paso: Revisar output y decidir mapeo a BD")


def main():
    # Path al Excel
    excel_path = Path(__file__).parent / "original-data" / "251028_Árbol web - control calidad.xlsx"

    if not excel_path.exists():
        print(f"❌ ERROR: Excel no encontrado en: {excel_path}")
        print("\nVerificar que el archivo existe en la carpeta original-data/")
        sys.exit(1)

    explore_excel(excel_path)


if __name__ == "__main__":
    main()
